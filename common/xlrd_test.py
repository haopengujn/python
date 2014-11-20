#-*- coding: utf8 -*-
#Filename xlrd_test.py


from openpyxl.reader.excel import load_workbook
import datetime

wb = load_workbook(filename = r'cdn.xlsx')
sheetnames = wb.get_sheet_names()  
ws = wb.get_sheet_by_name(sheetnames[1])
rowCnt = ws.get_highest_row()
columnCnt = ws.get_highest_column()

for row in range(1, 10):
    val=[]
    for column in range(columnCnt):
        cell=ws.cell(row=row, column=column).value
        if isinstance(cell, unicode):
            cell=cell.encode('utf8')
        if isinstance(cell, datetime.datetime):
            cell=cell.strftime("%Y-%m-%d")
        print cell
        